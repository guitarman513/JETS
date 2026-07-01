'''
here are the columns of the excel sheets
sometimes there is just CAT1, sometimes there is CAT1 and CAT2, sometimes there is cat1, cat2, and cat3. The code below will handle all of these cases.
CAT1	CAT2	CAT3	USELESSID	Description	Date	Mat. Unit	Material Price	Price Code	Labor Unit	Col 1 Labor	Col 2 Labor	Col 3 Labor	Col 4 Labor	Col 5 Labor	Col 6 Labor	Based On

each item will receive an auto-generated integer id

column description:
USELESSID: a useless id to be ignored
Description: the display name of the item
DATE: sometimes filled out. not sure format.
mat.unit: either E, C, or M for price unit of measure
material price: the price of the item. float.
price code: optional string. can be blank.
labor unit: either E, C, or M for labor unit of measure
col 1 labor: float. labor for labor category 1, etc for col2-6
based on: labor unit is based on this column. Either Len or Cnt.


overall flow:
read whole thing as a table in order
cats are only listed once. so keep same until see an update
rows with actual data in them will always have a value for "Based On"
if there is no data in "Based On", it means it isn't a row we care about.

Since items need a order_on_page and a parent_category_idx, let's make the hierarchy
for the categories first, then items

<NOCAT> MEANS NO CAT WHEN UPLOADING

'''
from typing import List,Dict
from models.database import Database, DatabaseItem, UnitMultiplier, LaborRates, UOM, DatabaseAssembly, AssemblyComponent, UILineItem
import openpyxl
from pathlib import Path

ITEM_FILE = Path.home() / "git" / "JETS" / "scripts" / "MANUAL-EXPORT-FROM-ACCUBID.xlsx"
EXPORT_FILE = Path.home() / "git" / "JETS" / "scripts" / "itemdb.json"

# Remember, all UILineItems have a parent_category_idx, but we don't store children
# So to get all TOP-LEVEL categories, we need to find all categories with parent_category_idx == None

ui_item_idx_counter:int = -1

db = Database(ui_dict={})


wb = openpyxl.load_workbook(ITEM_FILE)
all_sheets = wb.sheetnames

# loop thru sheets in wb
for sheetname in all_sheets:
    # WORKSHEET STUFF
    parent_category_idx_to_name:Dict[int,str] = {}
    parent_category_name_to_idx:Dict[str,int] = {}
    ws = wb[sheetname]
    colnames = [cell.value for cell in ws[1]]
    colname_to_idx = {colname: idx for idx, colname in enumerate(colnames)}
    colidx_to_colname = {idx: colname for idx, colname in enumerate(colnames)}
    current_cat1_name:str = None # type: ignore
    current_cat2_name:str = None # type: ignore
    current_cat3_name:str = None # type: ignore
    current_cat1_global_ui_index:int = None # type: ignore
    current_cat2_global_ui_index:int = None # type: ignore
    current_cat3_global_ui_index:int = None # type: ignore

    cat1_col:int = None # type: ignore
    cat2_col:int = None # type: ignore
    cat3_col:int = None # type: ignore
    is_cat2_present:bool = False
    is_cat3_present:bool = False

    # This is important. If we encounter a <NOCAT> in CAT2 or CAT3, then set this to a tier up
    line_item_parent_category_idx:int = None # type: ignore


    cat1_col:int = colname_to_idx.get("CAT1", -1)
    if "CAT2" in colnames:
        cat2_col = colname_to_idx.get("CAT2", -1)
        is_cat2_present = True
    if "CAT3" in colnames:
        cat3_col = colname_to_idx.get("CAT3", -1)
        is_cat3_present = True

    # LOOP THRU ROWS
    for idx, row in enumerate(ws.iter_rows()):
        if idx == 0:
            continue # skip header row
        if row[cat1_col].value and row[cat1_col].value != current_cat1_name:
            # then we need to save a new UILineItem Category for this new cat1
            ui_item_idx_counter += 1
            print(f"New CAT1: {row[cat1_col].value} at row {idx+1}. Assigning category_idx {ui_item_idx_counter}")
            current_cat1_name = str(row[cat1_col].value)
            current_cat1_global_ui_index = ui_item_idx_counter
            db.ui_dict[ui_item_idx_counter] = UILineItem(
                idx=ui_item_idx_counter,
                display_name=current_cat1_name,
                order_on_page=ui_item_idx_counter,
                parent_category_idx=None,
                line_item_type="CATEGORY"
            )

        if is_cat2_present and row[cat2_col].value and row[cat2_col].value != current_cat2_name:
            if row[cat2_col].value == "<NOCAT>":
                line_item_parent_category_idx = current_cat1_global_ui_index
            else:
                ui_item_idx_counter += 1
                print(f"New CAT2: {row[cat2_col].value} at row {idx+1}. Assigning category_idx {ui_item_idx_counter}")
                current_cat2_name = str(row[cat2_col].value)
                current_cat2_global_ui_index = ui_item_idx_counter
                line_item_parent_category_idx = current_cat2_global_ui_index
                db.ui_dict[ui_item_idx_counter] = UILineItem(
                    idx=ui_item_idx_counter,
                    line_item_type="CATEGORY",
                    display_name=current_cat2_name,
                    order_on_page=ui_item_idx_counter,
                    parent_category_idx=current_cat1_global_ui_index,
                )
        
        if is_cat3_present and row[cat3_col].value and row[cat3_col].value != current_cat3_name:
            if row[cat3_col].value == "<NOCAT>":
                line_item_parent_category_idx = current_cat2_global_ui_index
            else:
                ui_item_idx_counter += 1
                print(f"New CAT3: {row[cat3_col].value} at row {idx+1}. Assigning category_idx {ui_item_idx_counter}")
                current_cat3_name = str(row[cat3_col].value)
                current_cat3_global_ui_index = ui_item_idx_counter
                line_item_parent_category_idx = current_cat3_global_ui_index
                db.ui_dict[ui_item_idx_counter] = UILineItem(
                    idx=ui_item_idx_counter,
                    display_name=current_cat3_name,
                    order_on_page=ui_item_idx_counter,
                    parent_category_idx=current_cat2_global_ui_index,
                    line_item_type="CATEGORY",
                )

        item_desc = row[colname_to_idx.get("Description", -1)].value
        if item_desc and str(item_desc).strip() != "" and str(item_desc).strip() != "'":
            # then this is a row we care about. Create a DatabaseItem for it.
            ui_item_idx_counter += 1
            print(f"New ITEM: {item_desc} at row {idx+1}. Assigning item_idx {ui_item_idx_counter}")

            # determine UOM for price and labor
            price_uom_str = row[colname_to_idx.get("Mat. Unit", -1)].value
            labor_uom_str = row[colname_to_idx.get("Labor Unit", -1)].value
            price_uom = UOM.LENGTH if price_uom_str == "E" else UOM.COUNT if price_uom_str == "C" else UOM.LENGTH
            labor_uom = UOM.LENGTH if labor_uom_str == "E" else UOM.COUNT if labor_uom_str == "C" else UOM.LENGTH

            # determine price unit multiplier
            price_unit_multiplier = UnitMultiplier.E if price_uom_str == "E" else UnitMultiplier.C if price_uom_str == "C" else UnitMultiplier.M if price_uom_str == "M" else -1

            # determine labor unit multiplier
            labor_unit_multiplier = UnitMultiplier.E if labor_uom_str == "E" else UnitMultiplier.C if labor_uom_str == "C" else UnitMultiplier.M if labor_uom_str == "M" else -1

            # create LaborRates object
            labor_rates = LaborRates(
                labor1=row[colname_to_idx.get("Col 1 Labor", -1)].value or 0.0, #type: ignore
                labor2=row[colname_to_idx.get("Col 2 Labor", -1)].value or 0.0, #type: ignore
                labor3=row[colname_to_idx.get("Col 3 Labor", -1)].value or 0.0, #type: ignore
                labor4=row[colname_to_idx.get("Col 4 Labor", -1)].value or 0.0, #type: ignore
                labor5=row[colname_to_idx.get("Col 5 Labor", -1)].value or 0.0, #type: ignore
                labor6=row[colname_to_idx.get("Col 6 Labor", -1)].value or 0.0  #type: ignore
            )

            db.ui_dict[ui_item_idx_counter] = DatabaseItem(
                idx=ui_item_idx_counter,
                display_name=item_desc, #type: ignore
                order_on_page=ui_item_idx_counter,
                parent_category_idx=line_item_parent_category_idx,
                default_uom=price_uom,
                price_unit=price_unit_multiplier, #type: ignore
                price=row[colname_to_idx.get("Material Price", -1)].value or 0.0, #type: ignore
                labor_unit=labor_unit_multiplier, #type: ignore
                labor_rates=labor_rates,
                last_updated=row[colname_to_idx.get("Date", -1)].value.timestamp() if row[colname_to_idx.get("Date", -1)].is_date else 0.0, #type: ignore
                price_code=str(row[colname_to_idx.get("Price Code", -1)].value) or '',
                line_item_type="ITEM",
            )
        else:
            # then this is a row we don't care about. Skip it.
            print(f"Skipping row {idx+1} because it has no item description.")

    # end loop thru rows
# end loop thru sheets

print("ALL TOP LEVEL CATEGORIES WITH THEIR UI INDEX:")
print([f"{c.idx}: {c.display_name}" for c in db.ui_dict.values() if not c.parent_category_idx and c.line_item_type=="CATEGORY"])
print()

# Look for a random item and print its parent category hierarchy
lookupname = "CASHIER BOOTH PLC AUTOMATION PANEL"
some_item = next((item for item in db.ui_dict.values() if isinstance(item, DatabaseItem) and item.display_name == lookupname), None)
print(f"Path for item '{lookupname}':")
hierarchy:list[str] = []
if some_item:
    current_parent_idx = some_item.parent_category_idx
    while current_parent_idx is not None:
        parent_category = db.ui_dict.get(current_parent_idx)
        if parent_category and parent_category.line_item_type=="CATEGORY":
            hierarchy.append(parent_category.display_name)
            current_parent_idx = parent_category.parent_category_idx
        else:
            break
    hierarchy.reverse()
print(" -> ".join(hierarchy))

import random
print(f"four random db items: {random.sample(db.dbitems(), 4)}")


db.save_to_json_file(EXPORT_FILE)













'''
CAT1	CAT2	CAT3	USELESSID	Description	Date	Mat. Unit	Material Price	Price Code	Labor Unit	Col 1 Labor	Col 2 Labor	Col 3 Labor	Col 4 Labor	Col 5 Labor	Col 6 Labor	Based On
CONDUIT & FITTINGS	EMT	EMT CONDUIT	1	"  3/4"" EMT"	Tue Jun 23 2026 20:00:00 GMT-0400 (Eastern Daylight Time)	C	150	98001002002	C	3.52	4.4	5.28	5	6.2	7.5	Len
			3	"    1"" EMT"	Tue Jun 23 2026 20:00:00 GMT-0400 (Eastern Daylight Time)	C	180	98001002003	C	4.48	5.6	6.72	5.5	6.8	8.2	Len
			4	"1 1/4"" EMT"	Tue Jun 23 2026 20:00:00 GMT-0400 (Eastern Daylight Time)	C	275	98001002004	C	5.28	6.6	7.92	6.2	7.8	9.3	Len
			5	"1 1/2"" EMT"	Tue Jun 23 2026 20:00:00 GMT-0400 (Eastern Daylight Time)	C	350	98001002005	C	6.4	8	9.6	7	8.7	10.5	Len
			6	"    2"" EMT"	Tue Jun 23 2026 20:00:00 GMT-0400 (Eastern Daylight Time)	C	385	98001002006	C	8	10	12	8	10	12	Len
			7	"2 1/2"" EMT"	Tue Jun 23 2026 20:00:00 GMT-0400 (Eastern Daylight Time)	C	575	98001002007	C	9.92	12.4	14.88	9.5	11.8	14.2	Len
			8	"    3"" EMT"	Tue Jun 23 2026 20:00:00 GMT-0400 (Eastern Daylight Time)	C	735	98001002008	C	12.16	15.2	18.24	11	13.7	16.5	Len
			9	"3 1/2"" EMT"	Tue Jun 23 2026 20:00:00 GMT-0400 (Eastern Daylight Time)	C	1000	98001002009	C	13.5	16.87	20.25	13	16.2	19.5	Len
			10	"    4"" EMT"	Tue Jun 23 2026 20:00:00 GMT-0400 (Eastern Daylight Time)	C	1200	98001002010	C	14.72	18.4	22.08	16	20	24	Len
			11	"  1/2"" EMT"	Tue Jun 23 2026 20:00:00 GMT-0400 (Eastern Daylight Time)	C	100	98001002001	C	2.88	3.59	4.32	4.5	5.6	6.7	Len
			12													
		ELBOWS 	13	"  3/4"" EMT 90 ELBOW"	Tue Nov 18 2025 19:00:00 GMT-0500 (Eastern Standard Time)	C	666	98002002022	C	20	25	30	22	27	33	Cnt
			14	"    1"" EMT 90 ELBOW"	Tue Nov 18 2025 19:00:00 GMT-0500 (Eastern Standard Time)	C	800	98002002023	C	22	27.5	33	25	31	37	Cnt
			15	"1 1/4"" EMT 90 ELBOW"	Tue Nov 18 2025 19:00:00 GMT-0500 (Eastern Standard Time)	C	1000	98002002024	C	30	37.5	45	32	40	48	Cnt
			16	"1 1/2"" EMT 90 ELBOW"	Tue Nov 18 2025 19:00:00 GMT-0500 (Eastern Standard Time)	C	1300	98002002025	C	40	50	60	40	50	60	Cnt
'''
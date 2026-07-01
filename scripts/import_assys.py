'''
CAT1	CAT2	USELESSID	Description	Base	Fct 1	Fct 2	Catalog Number		
SWITCHES	SWITCHES  EMT	1	3way-30'CEILING-CONCRETE (new cond+wire each click)						
		1.1	  #12 THHN	Cnt	60	1			
		1.2	  R2  RED    WIRE CONN	Cnt	2	1			
		1.3	  '	Cnt	0	1			
		1.4	"  3/4"" EMT"	Cnt	30	1			

<NOCAT> MEANS NO CAT WHEN UPLOADING

'''
from typing import List,Dict
from models.database import Database, DatabaseItem, UOM, DatabaseAssembly, AssemblyComponent, UILineItem
import openpyxl
from pathlib import Path

ITEM_FILE = Path.home() / "git" / "JETS" / "scripts" / "EXPORT-ASSYS.xlsx"
ITEMDB_FILE = Path.home() / "git" / "JETS" / "scripts" / "itemdb.json"
EXPORT_FILE = Path.home() / "git" / "JETS" / "scripts" / "assydb.json"

# Remember, all UILineItems have a parent_category_idx, but we don't store children
# So to get all TOP-LEVEL categories, we need to find all categories with parent_category_idx == None

itemdb = Database.model_validate_json(ITEMDB_FILE.read_text())
itemdb_items = itemdb.dbitems()
itemdb_name_to_idx = {item.display_name.strip(): item.idx for item in itemdb_items}

ui_item_idx_counter:int = max(itemdb.ui_dict.keys())




wb = openpyxl.load_workbook(ITEM_FILE)
all_sheets = wb.sheetnames

'''
class UILineItem(BaseModel):
    idx: int
    display_name: str
    order_on_page: float
    parent_category_idx: int | None

class AssemblyComponent(BaseModel):
    kind: Literal["UISpacer", "DatabaseAssembly", "DatabaseItem"] # this is the type of the component. It can be a spacer, a category, or an item.
    # Below are only important if kind == "DatabaseItem" or "DatabaseAssembly"
    source_idx: int
    multiplier: float
    divisor: float
    uom: UOM

class DatabaseAssembly(UILineItem):
    components: list[AssemblyComponent]

assy_fragments = {
    ui_item_idx_counter:(parent_category_idx, display_name, [AssemblyComponent(...), AssemblyComponent(...), ...]),
}

'''
from pydantic import BaseModel
class AssemblyFragment(BaseModel):
    ui_item_idx_counter: int
    parent_category_idx: int
    assembly_name: str
    component: AssemblyComponent
# need to store assy fragments, then assemble later?
assembly_fragments:List[AssemblyFragment] = []

def cleanstr(s:str) -> str:
    return s.replace(" ", "").replace("-", "").replace("(", "").replace(")", "").replace(".", "").replace("[", "").replace("]", "").replace("'", "").replace('"','')


db = Database(ui_dict={})


# loop thru sheets in wb
for sheetname in all_sheets:
    # WORKSHEET STUFF
    parent_category_idx_to_name:Dict[int,str] = {}
    parent_category_name_to_idx:Dict[str,int] = {}
    ws = wb[sheetname]
    colnames = [cell.value for cell in ws[1]]
    colname_to_idx = {colname: idx for idx, colname in enumerate(colnames)}
    colidx_to_colname = {idx: colname for idx, colname in enumerate(colnames)}
    current_cat1_name:str = None # type: ignore
    current_cat2_name:str = None # type: ignore
    current_cat3_name:str = None # type: ignore
    current_cat1_global_ui_index:int = None # type: ignore
    current_cat2_global_ui_index:int = None # type: ignore
    current_cat3_global_ui_index:int = None # type: ignore

    cat1_col:int = None # type: ignore
    cat2_col:int = None # type: ignore
    cat3_col:int = None # type: ignore
    is_cat2_present:bool = False
    is_cat3_present:bool = False

    # This is important. If we encounter a <NOCAT> in CAT2 or CAT3, then set this to a tier up
    line_item_parent_category_idx:int = None # type: ignore


    cat1_col:int = colname_to_idx.get("CAT1", -1)
    if "CAT2" in colnames:
        cat2_col = colname_to_idx.get("CAT2", -1)
        is_cat2_present = True
    if "CAT3" in colnames:
        cat3_col = colname_to_idx.get("CAT3", -1)
        is_cat3_present = True

    # LOOP THRU ROWS
    current_assembly_name:str = None # type: ignore
    for idx, row in enumerate(ws.iter_rows()):
        item_desc = str(row[colname_to_idx.get("Description", -1)].value)

        if ui_item_idx_counter==7786:
            print("DEBUG:", [cell.value for cell in row])
        
        if "EMT NO WIRE" in str(item_desc):
            print("DEBUG:", [cell.value for cell in row])
        
        
        
        if idx == 0:
            continue # skip header row
        if row[cat1_col].value and row[cat1_col].value != current_cat1_name:
            # then we need to save a new UILineItem Category for this new cat1
            ui_item_idx_counter += 1
            print(f"New CAT1: {row[cat1_col].value} at row {idx+1}. Assigning category_idx {ui_item_idx_counter}")
            current_cat1_name = str(row[cat1_col].value)
            current_cat1_global_ui_index = ui_item_idx_counter
            line_item_parent_category_idx = current_cat1_global_ui_index
            db.ui_dict[ui_item_idx_counter] = UILineItem(
                idx=ui_item_idx_counter,
                display_name=current_cat1_name,
                order_on_page=ui_item_idx_counter,
                parent_category_idx=None,
                line_item_type="CATEGORY",
            )

        if is_cat2_present and row[cat2_col].value and row[cat2_col].value != current_cat2_name:
            if row[cat2_col].value == "<NOCAT>":
                line_item_parent_category_idx = current_cat1_global_ui_index
            else:
                ui_item_idx_counter += 1
                print(f"New CAT2: {row[cat2_col].value} at row {idx+1}. Assigning category_idx {ui_item_idx_counter}")
                current_cat2_name = str(row[cat2_col].value)
                current_cat2_global_ui_index = ui_item_idx_counter
                line_item_parent_category_idx = current_cat2_global_ui_index
                db.ui_dict[ui_item_idx_counter] = UILineItem(
                    idx=ui_item_idx_counter,
                    display_name=current_cat2_name,
                    line_item_type="CATEGORY",
                    order_on_page=ui_item_idx_counter,
                    parent_category_idx=current_cat1_global_ui_index,
                )
        
        if is_cat3_present and row[cat3_col].value and row[cat3_col].value != current_cat3_name:
            if row[cat3_col].value == "<NOCAT>":
                line_item_parent_category_idx = current_cat2_global_ui_index
            else:
                ui_item_idx_counter += 1
                print(f"New CAT3: {row[cat3_col].value} at row {idx+1}. Assigning category_idx {ui_item_idx_counter}")
                current_cat3_name = str(row[cat3_col].value)
                current_cat3_global_ui_index = ui_item_idx_counter
                line_item_parent_category_idx = current_cat3_global_ui_index
                db.ui_dict[ui_item_idx_counter] = UILineItem(
                    idx=ui_item_idx_counter,
                    line_item_type="CATEGORY",
                    display_name=current_cat3_name,
                    order_on_page=ui_item_idx_counter,
                    parent_category_idx=current_cat2_global_ui_index,
                )

        uselessid = row[colname_to_idx.get("USELESSID", -1)].value # whole numbers mark the start of an assy

        is_start_of_assy = int(uselessid)==float(uselessid) if uselessid is not None else False #type: ignore
        has_item_desc = item_desc is not None and str(item_desc).strip() != "" and str(item_desc).strip() != "'" #type: ignore

        if is_start_of_assy and has_item_desc:
            # then this is a row we care about. Create a DatabaseItem for it.
            ui_item_idx_counter += 1
            current_assembly_name = str(item_desc)
            print(f"New START OF ASSY: {current_assembly_name} at row {idx+1}. Assigning item_idx {ui_item_idx_counter}")

        elif '.' in str(row[colname_to_idx.get("USELESSID", -1)].value):
            print(f"New ASSY COMPONENT: {item_desc} at row {idx+1}. Assigning item_idx {ui_item_idx_counter} as part of assembly {current_assembly_name}")
            if "'" in str(row[colname_to_idx.get("Description", -1)].value):
                print("MARKING THIS AS A SPACER")
                assembly_fragments.append(
                    AssemblyFragment(
                        ui_item_idx_counter=ui_item_idx_counter,
                        parent_category_idx=line_item_parent_category_idx,
                        assembly_name=current_assembly_name,
                        component=AssemblyComponent(
                            line_item_type="SPACER",
                            source_idx=-1,
                            display_name=str(item_desc),
                            multiplier=1.0,
                            divisor=1.0,
                            uom=UOM.ABS,
                        )
                    )
                )
            else:
                _uom = row[colname_to_idx.get("Base", -1)].value
                # then this is a row we care about. Create an AssemblyComponent for it.
                sid=-1
                sid=itemdb_name_to_idx.get(item_desc.strip(), -1)
                if sid==-1:
                    print(f"WARNING: Could not find item '{item_desc}' in itemdb. It will be unlinked to the database.")
                    for item in itemdb_name_to_idx.keys():
                        _sid = itemdb_name_to_idx[item]
                        if cleanstr(item_desc) in cleanstr(item):
                            print(f"  Did you mean '{item}'? It is in the itemdb.")
                            resp = input("LINK TO THIS??? ")
                            if resp=='y':
                                sid=_sid
                                print(f"Assigning to {_sid}")

                    print()
                assembly_fragments.append(
                    AssemblyFragment(
                        ui_item_idx_counter=ui_item_idx_counter,
                        parent_category_idx=line_item_parent_category_idx,
                        assembly_name=current_assembly_name,
                        component=AssemblyComponent(
                            line_item_type="ITEM",
                            source_idx=sid,
                            display_name=str(item_desc),
                            multiplier=float(row[colname_to_idx.get("Fct 1", -1)].value), #type: ignore
                            divisor=float(row[colname_to_idx.get("Fct 2", -1)].value), #type: ignore
                            uom=UOM.COUNT if _uom=='Cnt' else UOM.LENGTH if _uom=='Len' else UOM.ABS if _uom=='Abs' else None, #type: ignore
                            )
                    )
                )

        else:
            print(f"Skipping row {idx+1}. Here is the row's data: {[cell.value for cell in row]}")

    # end loop thru rows
# end loop thru sheets


# Now construct assemblies

assys_with_unlinked_items:int=0
num_assemblies:int=0

for i in range(ui_item_idx_counter+1):
    # find all assembly fragments with this ui_item_idx_counter
    fragments_for_this_assembly = [frag for frag in assembly_fragments if frag.ui_item_idx_counter == i]
    if not fragments_for_this_assembly:
        continue
    
    totally_linked=True
    for f in fragments_for_this_assembly:
        if f.component.line_item_type == "ITEM" and f.component.source_idx == -1:
            totally_linked=False
            print(f"Assembly '{f.assembly_name}' with ui index {i} has an unlinked component: {f.component.display_name}")
    if not totally_linked:
        assys_with_unlinked_items += 1
    if "EMT NO WIRE" in fragments_for_this_assembly[0].assembly_name:
        print("DEBUG: ")
    
    if i==7583:
        print("DEBUG: ")
    
    
    # then we need to create a DatabaseAssembly for this ui_item_idx_counter
    db.ui_dict[i] = DatabaseAssembly(
        idx=i,
        display_name=fragments_for_this_assembly[0].assembly_name,
        order_on_page=i,
        parent_category_idx=fragments_for_this_assembly[0].parent_category_idx,
        components=[frag.component for frag in fragments_for_this_assembly],
        line_item_type="ASSEMBLY",
    )
    num_assemblies+=1
    print("Built Assembly named", fragments_for_this_assembly[0].assembly_name, "with ui index", i, "and parent_category_idx", fragments_for_this_assembly[0].parent_category_idx, )





print("ALL TOP LEVEL CATEGORIES WITH THEIR UI INDEX:")
print([f"{c.idx}: {c.display_name}" for c in db.ui_dict.values() if not c.parent_category_idx and c.line_item_type=="CATEGORY"])
print()

# Look for a random item and print its parent category hierarchy
lookupname = '''CAT 6 DATA W/ J-HOOKS 4 FT SPACING'''
some_item = next((item for item in db.ui_dict.values() if isinstance(item, DatabaseAssembly) and lookupname in item.display_name), None)
print(f"Path for item '{lookupname}':")
hierarchy:list[str] = []
if some_item:
    current_parent_idx = some_item.parent_category_idx
    while current_parent_idx is not None:
        parent_category = db.ui_dict.get(current_parent_idx)
        if parent_category and parent_category.line_item_type=="CATEGORY":
            hierarchy.append(parent_category.display_name)
            current_parent_idx = parent_category.parent_category_idx
        else:
            break
    hierarchy.reverse()
print(" -> ".join(hierarchy))


print("all_items that comprise this assembly:")
for item in some_item.components: #type:ignore
    if item.line_item_type == "ITEM":
        db_item = db.ui_dict.get(item.source_idx)
        if db_item and isinstance(db_item, DatabaseItem):
            print(f"  {db_item.display_name} (multiplier: {item.multiplier}, divisor: {item.divisor}, uom: {item.uom})")
        else:
            print(f"item {item.display_name} is unlinked to the database. source_idx: {item.source_idx}")
            print(f"it still has a multiplier, divisor, and uom though: {item.multiplier}, {item.divisor}, {item.uom}")
    elif item.line_item_type == "SPACER":
        print(f"  SPACER (multiplier: {item.multiplier}, divisor: {item.divisor}, uom: {item.uom})")


print()
print()

print(f"Number of assemblies with at least one unlinked component: {assys_with_unlinked_items}")
print(f"Total number of assemblies: {num_assemblies}")





db.save_to_json_file(EXPORT_FILE)



'''
CAT1	CAT2	USELESSID	Description	Base	Fct 1	Fct 2	Catalog Number		
SWITCHES	SWITCHES  EMT	1	3way-30'CEILING-CONCRETE (new cond+wire each click)						
		1.1	  #12 THHN	Cnt	60	1			
		1.2	  R2  RED    WIRE CONN	Cnt	2	1			
		1.3	  '	Cnt	0	1			
		1.4	"  3/4"" EMT"	Cnt	30	1			
'''